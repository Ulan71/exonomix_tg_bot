import pdfplumber
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl import load_workbook


# pd.set_option('display.max_rows', None)
# pd.set_option('display.max_columns', None)
# pd.set_option('display.width', None)
# pd.set_option('display.max_colwidth', None)

pdf_path = "pdf_files"
excel_path = "sheet.xlsx"


class PDFFileProcessing:
    def __init__(self, path, excel_path, uuid):
        self.path = path
        self.excel_path = excel_path
        self.uuid = uuid

        self.progress = 0

    def files_processing(self):
        files_list = os.listdir(self.path)

        wb = load_workbook(self.excel_path)
        sheet = wb.active
        sheet.insert_cols(8)

        pdf_data = {}

        # print(os.listdir(r"pdf_files"))

        for file_num, file in enumerate(files_list):
            # print(file)
            pdf_path = rf"{self.path}/{file}"

            gws = None
            fpn = None  # found page num
            gdfs = []  # got data from sheet
            pdf_esf = None
            pl = 0

            with pdfplumber.open(pdf_path) as pdf:
                file_length = len(pdf.pages)  # количество страниц в файле

                for page_num in range(file_length):  # Перебор всех страниц
                    page = pdf.pages[page_num]  # Первый этап парсинга страницы

                    # print(page)
                    table = page.extract_table()

                    try:
                        for row in table:
                            for cell in row:
                                if "Регистрационный номер ESF" in str(cell):
                                    pdf_esf = cell[-34:]
                                    raise TypeError
                    except:
                        pass

                    if table == None:
                        continue

                    df = pd.DataFrame(table[1:], columns=table[0])

                    row, col = df.shape

                    for c in range(col):
                        for r in range(row):
                            text = str(df.iloc[r, c]).replace("\n", " ")

                            if "Наименование товаров, работ, услуг" in text:
                                gws = c
                                fpn = page_num
                                # print("FOUND", page_num, c, file)

        #  <=================================================================================>

                file_length = len(pdf.pages)  # количество страниц в файле

                try:
                    for page_num2 in range(fpn, file_length):  # Перебор всех страниц
                        page22 = pdf.pages[page_num2]  # Первый этап парсинга страницы

                        table2 = page22.extract_table()

                        if table2 == None:
                            continue

                        df2 = pd.DataFrame(table2[1:], columns=table2[0])
                        row2, col2 = df2.shape

                        if gws > col2:
                            continue

                        data_tabel = df2.iloc[:, gws]

                        for data in data_tabel:
                            d_type = type(data)
                            if data != None:
                                if d_type != float and data != "nan":
                                    if not data.isdigit():
                                        cleaned_data = data.replace("\n", " ")
                                        if not cleaned_data == "Наименование товаров, работ, услуг":
                                            gdfs.append(cleaned_data)

                except Exception as ex:
                    # print(ex)
                    # print(file)
                    pass

            # pdf_data[pdf_esf] = gdfs
            print(pdf_esf, gdfs, pdf_path)

            for num, esf_cell in enumerate(sheet["G"]):
                if pdf_esf in str(esf_cell.value):
                    if not len(gdfs) == 0:
                        sheet[f"H{num+1}"] = gdfs[0]

                    sheet.cell(row=num+1, column=7).hyperlink = pdf_path
                    sheet.cell(row=num+1, column=7).style = "Hyperlink"

                    print("found")

                    self.get_file_progress(len(files_list), file_num)
                    # wb.save("sheet.xlsx")

        wb.save(excel_path)

    def get_file_progress(self, files_count, current_file):
        self.progress = round((current_file * 100) * files_count, 2)


x = PDFFileProcessing(pdf_path, excel_path, "1111")
x.files_processing()


