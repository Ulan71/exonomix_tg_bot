import time
from builtins import print

from flask import Flask, render_template, request, jsonify, redirect, send_file
import shutil
import os
import pdfplumber
import uuid
import pdfplumber
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import threading
from waitress import serve
from werkzeug.serving import run_simple


# pd.set_option('display.max_rows', None)
# pd.set_option('display.max_columns', None)
# pd.set_option('display.width', None)
# pd.set_option('display.max_colwidth', None)


class PDFFileProcessing:
    def __init__(self, path, uuid):
        self.path = path
        self.excel_path = f"{UPLOAD_FOLDER}/{uuid}/sheet.xlsx"
        self.uuid = uuid

        self.progress = 0
        self.current_file_name = ''

    def combine_excel_files(self):
        try:
            xlsx_files = []

            for file in os.listdir(self.path):  # Get all files from path
                name, ext = os.path.splitext(file)  # Split text to get ext
                if ext == ".xlsx":
                    xlsx_files.append(file)  # Add only xlsx files

            xlsx_file_row = []  # List for xlsx files rows, we need it for copy values from book

            if not len(xlsx_files) == 1:  # If we have only one file we don't start copy
                for xlsx_file in xlsx_files[1:]:  # Working with all xlsx files, except the first one
                    wb = load_workbook(f"{self.path}/{xlsx_file}")  # Just load the book
                    sheet = wb.active  # Activating current sheet

                    for row in sheet.iter_rows():  # Reading rows
                        row_data = []  # List for temp values

                        for cell in row:  # Reading each cell
                            row_data.append(cell.value)  # Add cell value to list

                        xlsx_file_row.append(row_data)  # Add row list form 'row_data' to 'xlsx_file_row'

                none_counter = 0  # Counter of None

                for last_row_value in xlsx_file_row[-1]:  # Get last row
                    if last_row_value is None:  # If cell from last row equal None
                        none_counter += 1  # +1 to 'none_counter'

                if none_counter == len(xlsx_file_row[-1]):  # If 'none_counter' equal 'xlsx_file_row[-1]' len
                    del xlsx_file_row[-1]  # del

                del xlsx_file_row[:2]

                wb = load_workbook(f"{self.path}/{xlsx_files[0]}")  # Load main Excel file
                ws = wb.active  # Activate book

                none_counter = 0  # None counter

                for last_row_value in ws.iter_rows(min_row=int(ws.max_row),
                                                   max_row=int(ws.max_row)):  # Get the last row values
                    for cell in last_row_value:  # Get each cell value
                        if cell.value is None:  # If we have None, +1 to 'none_counter'
                            none_counter += 1

                if none_counter == len(
                        last_row_value):  # If 'none_counter' and 'last_row_vlaue' len equal, del the last row
                    ws.delete_rows(int(ws.max_row))

                for row in xlsx_file_row:  # Adding data from books
                    ws.append(row)

                wb.save(f"{self.path}/sheet.xlsx")  # Save all changes
                # print(self.path, "<===============")

                for book in xlsx_files[1:]: os.remove(f"{self.path}/{book}")  # Delete books

            return True
        except Exception as e:
            return False, str(e)

    def files_processing(self):
        os.makedirs(f"{UPLOAD_FOLDER}/{self.uuid}_ready", exist_ok=True)
        os.makedirs(f"{UPLOAD_FOLDER}/{self.uuid}_ready/PDFs", exist_ok=True)
        files_list = [
            filename
            for filename in os.listdir(self.path)
            if filename.lower().endswith('.pdf')
        ]

        wb = load_workbook(self.excel_path)
        sheet = wb.active
        sheet.insert_cols(8)

        # print(os.listdir(r"pdf_files"))

        for file_num, file in enumerate(files_list):

            # print(file)
            pdf_path = rf"{self.path}/{file}"

            gws = None
            fpn = None  # found page num
            gdfs = []  # got data from sheet
            pdf_esf = None
            pl = 0

            with pdfplumber.open(pdf_path) as pdf:
                file_length = len(pdf.pages)  # количество страниц в файле

                for page_num in range(file_length):  # Перебор всех страниц
                    page = pdf.pages[page_num]  # Первый этап парсинга страницы

                    # print(page)
                    table = page.extract_table()

                    try:
                        for row in table:
                            for cell in row:
                                if "Регистрационный номер ESF" in str(cell):
                                    pdf_esf = cell[-34:]
                                    raise TypeError
                    except:
                        pass

                    if table == None:
                        continue

                    df = pd.DataFrame(table[1:], columns=table[0])

                    row, col = df.shape

                    for c in range(col):
                        for r in range(row):
                            text = str(df.iloc[r, c]).replace("\n", " ")

                            if "Наименование товаров, работ, услуг" in text:
                                gws = c
                                fpn = page_num
                                # print("FOUND", page_num, c, file)

                #  <=================================================================================>

                file_length = len(pdf.pages)  # количество страниц в файле

                try:
                    for page_num2 in range(fpn, file_length):  # Перебор всех страниц
                        page22 = pdf.pages[page_num2]  # Первый этап парсинга страницы

                        table2 = page22.extract_table()

                        if table2 == None:
                            continue

                        df2 = pd.DataFrame(table2[1:], columns=table2[0])
                        row2, col2 = df2.shape

                        if gws > col2:
                            continue

                        data_tabel = df2.iloc[:, gws]

                        for data in data_tabel:
                            d_type = type(data)
                            if data != None:
                                if d_type != float and data != "nan":
                                    if not data.isdigit():
                                        cleaned_data = data.replace("\n", " ")
                                        if not cleaned_data == "Наименование товаров, работ, услуг":
                                            gdfs.append(cleaned_data)

                except Exception as ex:
                    # print(ex)
                    # print(file)
                    pass

            # pdf_data[pdf_esf] = gdfs
            # print(pdf_esf, gdfs, pdf_path)

            for num, esf_cell in enumerate(sheet["G"]):
                if pdf_esf in str(esf_cell.value):
                    if not len(gdfs) == 0:
                        sheet[f"H{num + 1}"] = gdfs[0]  # Insert data to Excel sheet
                        # print(gdfs)

                    if len(gdfs) == 0:
                        fgdfs = "None"
                    else:
                        fgdfs = str(gdfs[0])

                    for symbol in ["<", ">", ":", "|", "/", "?", "\\", '"', "*", "\n", "  "]:
                        if symbol in fgdfs:
                            fgdfs = fgdfs.replace(symbol, " ")

                    fgdfs = fgdfs[:80]
                    # print(fgdfs)

                    sheet.cell(row=num + 1, column=7).hyperlink = f"PDFs/{file_num + 1} {fgdfs}.pdf"
                    sheet.cell(row=num + 1, column=7).style = "Hyperlink"

                    # print("found")
                    # print(file)
                    shutil.copy(f"{UPLOAD_FOLDER}/{self.uuid}/{file}",
                                f"{UPLOAD_FOLDER}/{self.uuid}_ready/PDFs/{file_num + 1} {fgdfs}.pdf")

                    self.get_file_progress(len(files_list), file_num, file)

        wb.save(self.excel_path)

        shutil.copy(f"{self.excel_path}", f"{UPLOAD_FOLDER}/{self.uuid}_ready")

        # shutil.make_archive(f'{UPLOAD_FOLDER}/{self.uuid}', 'zip', f'{UPLOAD_FOLDER}/{self.uuid}_ready')

    def start_processing(self):
        combine = self.combine_excel_files()
        print(combine)

        if combine:
            self.files_processing()

    def get_file_progress(self, files_count, current_file, file_name):
        self.progress = round((current_file * 100) / (files_count - 1), 2)
        self.current_file_name = file_name


app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
app.config['MAX_FORM_PARTS'] = 10000
app.config['MAX_CONTENT_LENGTH'] = 2 * 1024 * 1024 * 1024
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
tasks = {}


@app.route("/")
def hello_world():
    return render_template("index.html")


@app.route('/upload', methods=['POST'])
def upload_file():
    session_uuid = str(uuid.uuid4())

    files = request.files.getlist("files")
    print(files)

    os.makedirs(f"{UPLOAD_FOLDER}/{session_uuid}", exist_ok=True)

    for file in files:
        filename = file.filename
        file.save(f"{UPLOAD_FOLDER}/{session_uuid}/{filename}")
        print(filename)

    global tasks

    tasks[session_uuid] = PDFFileProcessing(f"{UPLOAD_FOLDER}/{session_uuid}", session_uuid)
    thread = threading.Thread(target=tasks[session_uuid].start_processing)
    thread.start()

    return redirect(f"/result/{session_uuid}")


@app.route('/result/<session_uuid>', methods=['GET'])
def result_page(session_uuid):
    return render_template("result.html", session_uuid=session_uuid)


@app.route('/process/<session_uuid>', methods=['GET'])
def get_process(session_uuid):
    global tasks
    return jsonify({
        "progress": tasks[session_uuid].progress,
        "current_file": tasks[session_uuid].current_file_name
    })


@app.route('/files/<session_uuid>', methods=['GET'])
def get_user_files(session_uuid):
    session_folder = os.path.join(UPLOAD_FOLDER, session_uuid)

    if not os.path.exists(session_folder):
        return jsonify({
            "error": "Session not found"
        }), 404

    files = os.listdir(session_folder)

    return jsonify({
        "session_id": session_uuid,
        "files": files
    })


@app.route('/download/<session_uuid>', methods=['GET'])
def download_file(session_uuid):
    ready_archive = shutil.make_archive(f'{UPLOAD_FOLDER}/{session_uuid}', 'zip',
                                        f'{UPLOAD_FOLDER}/{session_uuid}_ready')
    return send_file(ready_archive, as_attachment=True)


if __name__ == "__main__":
    # app.run(debug=True)
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=True, threaded=True)
    # serve(app, host='0.0.0.0', port=5000)
